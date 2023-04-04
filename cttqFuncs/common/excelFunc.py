'''
Author: Logic
Date: 2022-09-08 11:03:51
LastEditTime: 2022-09-26 19:33:23
Description: 
'''
from typing import Any, List, Dict
import xlwt
import xlrd2
import openpyxl
from ..basic.exClass import CommonException
from ..basic.exFunc import *
from .fileFunc import readLines, isExist, createFile, deleteFile, writeAppend


class XlsWriter():

    def __init__(self, path: str) -> None:
        self.path = path

        if path.endswith('.xlsx'):
            self.isXlsx = True
            self.workbook = openpyxl.Workbook(write_only=True)
        elif path.endswith('.xls'):
            self.isXlsx = False
            self.workbook = xlwt.Workbook(encoding="utf-8")
        else:
            raise CommonException('文件类型不正确')

    def save(self):
        self.workbook.save(self.path)

    def createSheet(self, headers: List[str], datas: List[List[Any]],
                    sheetName: str='sheet1'):
        # 生成sheet
        if self.isXlsx:

            sheet = self.workbook.create_sheet(sheetName)
            # 写入标题
            sheet.append(headers)

            # 写入每一行
            for data in datas:
                sheet.append(data)
        else:

            sheet = self.workbook.add_sheet(sheetName)
            # 写入标题
            for col, column in enumerate(headers):
                sheet.write(0, col, column)

            # 写入每一行
            for row, data in enumerate(datas):
                for col, col_data in enumerate(data):
                    sheet.write(row + 1, col, col_data)

    def createSheetWithDict(self, dataList: List[Dict[str, Any]],
                            sheetName: str='sheet1'):
        if len(dataList) > 10:
            headers = list(set(dataList[0:10].flatMap(lambda d: d.ks())))
        else:
            headers = list(set(dataList.flatMap(lambda d: d.ks())))
        datas = []
        for data in dataList:
            tmp = []
            for t in headers:
                if t in data and data[t]:
                    tmp.append(data[t])
                else:
                    tmp.append('')
            datas.append(tmp)
        self.createSheet(headers, datas, sheetName)


class XlsReader():

    def __init__(self, path) -> None:
        self.workbook = xlrd2.open_workbook(path)

    def sheetNames(self):
        return self.workbook.sheet_names()

    def readSheetByName(self, sheetName: str) -> List[Dict[str, Any]]:
        sh = self.workbook.sheet_by_name(sheetName)
        re = []
        headers = sh.row_values(0)
        for i in range(1, sh.nrows):
            d = dict(zip(headers, sh.row_values(i)))
            re.append(d)
        return re

    def readSheetByIndex(self, sheetIndex: int=0) -> List[Dict[str, Any]]:
        sh = self.workbook.sheet_by_index(sheetIndex)
        re = []
        headers = sh.row_values(0)
        for i in range(1, sh.nrows):
            d = dict(zip(headers, sh.row_values(i)))
            re.append(d)
        return re

    def readSheetAsMatrixByName(self, sheetName: str) -> List[Dict[str, Any]]:
        sh = self.workbook.sheet_by_name(sheetName)
        re = []
        for i in range(0, sh.nrows):
            re.append(sh.row_values(i))
        return re


class CsvRW():
    def __init__(self, path, splitFlag: str = ',', encoding='utf-8') -> None:
        self.path = path
        self.splitFlag = splitFlag
        self.encoding = encoding

        if not isExist(self.path):
            createFile(self.path)

    def headers(self):
        lines = readLines(self.path, 1, encoding=self.encoding)
        if len(lines) >= 1:
            return lines[0].split(self.splitFlag).filter(lambda s: s != '')
        return []

    def read(self):
        lines = readLines(self.path, encoding=self.encoding)
        re = []
        headers = self.headers()
        for line in lines:
            d = dict(zip(headers, line.split(self.splitFlag)))
            re.append(d)
        return re

    def clear(self):
        deleteFile(self.path)
        createFile(self.path)

    def write(self, datas: List[dict]):
        headers = self.headers()
        strs = []
        if len(headers) == 0:
            headers = list(datas[0].keys())
            strs.append(self.splitFlag.join(headers)+'\n')

        for data in datas:
            tmp = []
            for t in headers:
                if t in data and data[t]:
                    tmp.append(
                        '"'+str(data[t]).replace('\\', '\\\\').replace('"', '\"')+'"')
                else:
                    tmp.append('""')
            strs.append(self.splitFlag.join(tmp)+'\n')
        writeAppend(self.path, lines=strs, encoding=self.encoding)
